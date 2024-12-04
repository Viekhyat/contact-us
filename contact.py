from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os

app = Flask(__name__)

# Path to Excel file
file_path = r"C:\Users\viekh\OneDrive\Desktop\.vscode\self airline\contact_us\contact-us.xlsx"

# Function to store data in Excel
def save_to_excel(name, email, message):
    # Check if file exists
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Email", "Message"])  # Add headers if file is newly created

    # Add data to the next row
    sheet.append([name, email, message])
    workbook.save(file_path)

# Function to send acknowledgment email
def send_acknowledgment_email(receiver_email):
    sender_email = os.getenv("SENDER_EMAIL", 'avajava0802@gmail.com')
    sender_password = os.getenv("SENDER_PASSWORD", 'viekhyatarpita')

    subject = "Skyhulk Airlines - Contact Confirmation"
    body = "Thank you for reaching out! We will contact you soon."

    # Email setup
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receiver_email
    message['Subject'] = subject
    message.attach(MIMEText(body, 'plain'))

    # Send the email
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, receiver_email, message.as_string())
    except Exception as e:
        print("Error sending email:", e)

# Route for form submission
@app.route('/submit-form', methods=['POST'])
def submit_form():
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    message = data.get('message')

    # Save to Excel
    save_to_excel(name, email, message)

    # Send acknowledgment email
    send_acknowledgment_email(email)

    return jsonify({"status": "success", "message": "Form submitted successfully"})

if __name__ == '__main__':
    app.run(debug=True)
